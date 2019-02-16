import requests
import base64
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = 'Backout'

response = requests.get('https://jsonplaceholder.typicode.com/users')
cabecalho = ['id', 'nome', 'email', 'telefone']
for value in cabecalho:
    ws.cell(row=1, column=cabecalho.index(value) + 1, value=value)
columns = len(cabecalho)
if response.status_code == 200:
    list = response.json()
    rows = len(list)
    for i, item in enumerate(list, 1):
        ws.cell(row=i, column=1, value=item['id'])
        ws.cell(row=i, column=2, value=item['name'])
        ws.cell(row=i, column=3, value=item['email'])
        ws.cell(row=i, column=4, value=item['phone'])
    wb.save('backout.xlsx')
