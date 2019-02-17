import requests
from openpyxl import Workbook

response  = requests.get("https://jsonplaceholder.typicode.com/photos")

wb = Workbook()
ws = wb.active
ws.title = 'Backout'

def criar_cabecalho(json, ws):
    for col, name in  enumerate(json[0].keys(),1):
        ws.cell(row=1, column= col, value=name)

def criar_registros(json, ws):
    for row, item in enumerate(json,2):
        for column, key in enumerate(json[0].keys(),1):
            try:
                ws.cell(row=row, column= column, value=item[key])
            except ValueError as identifier:
                pass


if response.status_code == 200:
    json = response.json()
    criar_cabecalho(json, ws)
    criar_registros(json, ws)
    wb.save("fotos.xlsx")